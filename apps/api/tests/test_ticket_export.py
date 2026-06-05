import io
import uuid
from datetime import UTC, datetime
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from star_itsm_api.core.security import ROLE_SUBMITTER, get_current_user
from star_itsm_api.main import app
from star_itsm_api.services import ticket_export as export_service


def _fake_end_user() -> SimpleNamespace:
    return SimpleNamespace(
        id="00000000-0000-0000-0000-000000000099",
        email="submitter@example.dk",
        display_name="Submitter",
        role=ROLE_SUBMITTER,
        is_active=True,
        password_hash=None,
        deleted_at=None,
        organization_id=None,
    )


@pytest.mark.asyncio
async def test_tickets_export_forbidden_for_end_user(client: AsyncClient) -> None:
    app.dependency_overrides[get_current_user] = _fake_end_user
    try:
        response = await client.get("/api/v1/reports/tickets/export")
        assert response.status_code == 403
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.mark.asyncio
async def test_tickets_export_without_database_returns_503(client: AsyncClient) -> None:
    response = await client.get("/api/v1/reports/tickets/export")
    assert response.status_code == 503


def _export_ticket(**kwargs: object) -> SimpleNamespace:
    now = datetime(2026, 6, 1, 9, 0, tzinfo=UTC)
    defaults: dict[str, object] = {
        "id": uuid.uuid4(),
        "ticket_number": "INC-2026-00001",
        "title": "Printer virker ikke",
        "organization_id": uuid.uuid4(),
        "status": "new",
        "priority": "high",
        "ticket_type": "incident",
        "is_major": True,
        "is_shared": True,
        "created_at": now,
        "updated_at": now,
        "resolved_at": now,
        "closed_at": now,
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


@pytest.mark.asyncio
async def test_load_organization_names_empty_returns_empty() -> None:
    db = AsyncMock()
    result = await export_service._load_organization_names(db, set())
    assert result == {}
    db.execute.assert_not_called()


@pytest.mark.asyncio
async def test_load_organization_names_returns_mapping() -> None:
    org_id = uuid.uuid4()
    org = SimpleNamespace(id=org_id, name="Stjernekommune")
    execute_result = MagicMock()
    execute_result.scalars.return_value.all.return_value = [org]
    db = AsyncMock()
    db.execute = AsyncMock(return_value=execute_result)

    result = await export_service._load_organization_names(db, {org_id})
    assert result == {org_id: "Stjernekommune"}


@pytest.mark.asyncio
async def test_fetch_tickets_for_export_agent_applies_team_filter() -> None:
    ticket = _export_ticket()
    execute_result = MagicMock()
    execute_result.scalars.return_value.all.return_value = [ticket]
    db = AsyncMock()
    db.execute = AsyncMock(return_value=execute_result)
    user = SimpleNamespace(role="agent", organization_id=None)

    with (
        patch.object(export_service, "apply_ticket_list_filter", side_effect=lambda stmt, u: stmt),
        patch.object(export_service, "is_staff_role", return_value=True),
        patch.object(
            export_service,
            "apply_agent_team_list_filter",
            AsyncMock(side_effect=lambda db, stmt, u: stmt),
        ) as mock_team_filter,
    ):
        tickets = await export_service.fetch_tickets_for_export(db, user)

    assert tickets == [ticket]
    mock_team_filter.assert_awaited_once()


@pytest.mark.asyncio
async def test_fetch_tickets_for_export_admin_skips_team_filter() -> None:
    ticket = _export_ticket()
    execute_result = MagicMock()
    execute_result.scalars.return_value.all.return_value = [ticket]
    db = AsyncMock()
    db.execute = AsyncMock(return_value=execute_result)
    user = SimpleNamespace(role="admin", organization_id=None)

    with (
        patch.object(export_service, "apply_ticket_list_filter", side_effect=lambda stmt, u: stmt),
        patch.object(export_service, "is_staff_role", return_value=True),
        patch.object(
            export_service,
            "apply_agent_team_list_filter",
            AsyncMock(side_effect=lambda db, stmt, u: stmt),
        ) as mock_team_filter,
    ):
        tickets = await export_service.fetch_tickets_for_export(db, user)

    assert tickets == [ticket]
    mock_team_filter.assert_not_called()


@pytest.mark.asyncio
async def test_build_tickets_export_xlsx_full_rows() -> None:
    enriched_ticket = _export_ticket()
    # Second ticket has no enriched row and no organization -> exercises else branches.
    plain_ticket = _export_ticket(
        id=uuid.uuid4(),
        ticket_number="INC-2026-00002",
        organization_id=None,
        is_major=False,
        is_shared=False,
        priority="weird",
        ticket_type="weird",
        status="weird",
        created_at=None,
        updated_at=None,
        resolved_at=None,
        closed_at=None,
    )
    tickets = [enriched_ticket, plain_ticket]

    enriched_row = SimpleNamespace(
        id=enriched_ticket.id,
        assigned_user_name="Anna Agent",
        assigned_team_name="Servicedesk",
        reporter_display_name="Borger Hansen",
    )

    org_id = enriched_ticket.organization_id
    org = SimpleNamespace(id=org_id, name="Stjernekommune")
    org_result = MagicMock()
    org_result.scalars.return_value.all.return_value = [org]
    db = AsyncMock()
    db.execute = AsyncMock(return_value=org_result)

    with (
        patch.object(
            export_service,
            "fetch_tickets_for_export",
            AsyncMock(return_value=tickets),
        ),
        patch.object(
            export_service,
            "tickets_to_read_list",
            AsyncMock(return_value=[enriched_row]),
        ),
    ):
        data = await export_service.build_tickets_export_xlsx(db, MagicMock())

    assert isinstance(data, bytes)
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["Sager"]
    assert sheet.cell(row=1, column=1).value == "Sagsnr"
    # Row 2 = enriched ticket
    assert sheet.cell(row=2, column=1).value == "INC-2026-00001"
    assert sheet.cell(row=2, column=3).value == "Stjernekommune"
    assert sheet.cell(row=2, column=4).value == "Ny"
    assert sheet.cell(row=2, column=5).value == "Høj"
    assert sheet.cell(row=2, column=6).value == "Hændelse"
    assert sheet.cell(row=2, column=7).value == "Anna Agent"
    assert sheet.cell(row=2, column=8).value == "Servicedesk"
    assert sheet.cell(row=2, column=9).value == "Borger Hansen"
    assert sheet.cell(row=2, column=10).value == "Ja"
    assert sheet.cell(row=2, column=11).value == "Ja"
    # Row 3 = plain ticket fallbacks
    assert sheet.cell(row=3, column=3).value in (None, "")
    assert sheet.cell(row=3, column=5).value == "weird"
    assert sheet.cell(row=3, column=7).value in (None, "")
    assert sheet.cell(row=3, column=10).value == "Nej"
    assert sheet.cell(row=3, column=11).value == "Nej"
    # Footer
    footer_row = len(tickets) + 3
    assert "Eksporteret" in sheet.cell(row=footer_row, column=1).value
    assert sheet.cell(row=footer_row + 1, column=1).value == "Antal sager: 2"


@pytest.mark.asyncio
async def test_build_tickets_export_xlsx_empty() -> None:
    db = AsyncMock()
    db.execute = AsyncMock()

    with (
        patch.object(
            export_service,
            "fetch_tickets_for_export",
            AsyncMock(return_value=[]),
        ),
        patch.object(
            export_service,
            "tickets_to_read_list",
            AsyncMock(return_value=[]),
        ),
    ):
        data = await export_service.build_tickets_export_xlsx(db, MagicMock())

    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["Sager"]
    # No tickets -> footer_row = 0 + 3 = 3 (Eksporteret), count on row 4.
    assert "Eksporteret" in sheet.cell(row=3, column=1).value
    assert sheet.cell(row=4, column=1).value == "Antal sager: 0"
