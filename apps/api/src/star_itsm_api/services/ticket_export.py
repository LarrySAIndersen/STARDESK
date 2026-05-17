import io
import uuid
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from star_itsm_api.models.organization import Organization
from star_itsm_api.models.ticket import Ticket
from star_itsm_api.models.user import User
from star_itsm_api.services.org_access import (
    apply_agent_team_list_filter,
    apply_ticket_list_filter,
)
from star_itsm_api.services.permissions import is_staff_role
from star_itsm_api.services.reports import status_label_da
from star_itsm_api.services.ticket_read import tickets_to_read_list

EXPORT_HEADERS = [
    "Sagsnr",
    "Titel",
    "Organisation",
    "Status",
    "Prioritet",
    "Type",
    "Sagsbehandler",
    "Gruppe",
    "Indmelder",
    "Stor sag",
    "Delt sag",
    "Oprettet",
    "Opdateret",
    "Løst",
    "Lukket",
]

PRIORITY_LABELS_DA: dict[str, str] = {
    "critical": "Kritisk",
    "high": "Høj",
    "medium": "Medium",
    "low": "Lav",
}

TICKET_TYPE_LABELS_DA: dict[str, str] = {
    "service_request": "Serviceanmodning",
    "incident": "Hændelse",
    "problem": "Problem",
}


async def _load_organization_names(
    db: AsyncSession,
    org_ids: set[uuid.UUID],
) -> dict[uuid.UUID, str]:
    if not org_ids:
        return {}
    rows = await db.execute(select(Organization).where(Organization.id.in_(org_ids)))
    return {org.id: org.name for org in rows.scalars().all()}


async def fetch_tickets_for_export(db: AsyncSession, user: User) -> list[Ticket]:
    stmt = select(Ticket).where(Ticket.deleted_at.is_(None))
    stmt = apply_ticket_list_filter(stmt, user)
    if is_staff_role(user) and user.role == "agent":
        stmt = await apply_agent_team_list_filter(db, stmt, user)
    stmt = stmt.order_by(Ticket.created_at.desc())
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def build_tickets_export_xlsx(db: AsyncSession, user: User) -> bytes:
    tickets = await fetch_tickets_for_export(db, user)
    enriched = await tickets_to_read_list(db, tickets)
    enrich_by_id = {t.id: t for t in enriched}

    org_ids = {t.organization_id for t in tickets if t.organization_id}
    org_names = await _load_organization_names(db, org_ids)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sager"

    header_font = Font(bold=True)
    for col, header in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font

    for row_idx, ticket in enumerate(tickets, start=2):
        row = enrich_by_id.get(ticket.id)
        org_name = org_names.get(ticket.organization_id) if ticket.organization_id else ""
        values = [
            ticket.ticket_number,
            ticket.title,
            org_name,
            status_label_da(ticket.status),
            PRIORITY_LABELS_DA.get(ticket.priority, ticket.priority),
            TICKET_TYPE_LABELS_DA.get(ticket.ticket_type, ticket.ticket_type),
            row.assigned_user_name if row else "",
            row.assigned_team_name if row else "",
            row.reporter_display_name if row else "",
            "Ja" if ticket.is_major else "Nej",
            "Ja" if getattr(ticket, "is_shared", False) else "Nej",
            ticket.created_at.isoformat() if ticket.created_at else "",
            ticket.updated_at.isoformat() if ticket.updated_at else "",
            ticket.resolved_at.isoformat() if ticket.resolved_at else "",
            ticket.closed_at.isoformat() if ticket.closed_at else "",
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_idx, column=col, value=value)

    footer_row = len(tickets) + 3
    sheet.cell(row=footer_row, column=1, value=f"Eksporteret {datetime.now(UTC).isoformat()}")
    sheet.cell(row=footer_row + 1, column=1, value=f"Antal sager: {len(tickets)}")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
